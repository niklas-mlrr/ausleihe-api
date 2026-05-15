#!/usr/bin/env python3
"""Aktualisiert die 'Bestand'-Zellen in der Excel-Datei mit Daten aus der Ausleihe-API.

Verwendung:
    python3 "bestand- und nachbestellungen/update_bestand.py"

Die ISBN-zu-Zelle-Zuordnung wird aus config.json gelesen (aus dem alten Webscraper).
Aktualisiert werden nur die 'Bestand'-Zellen (Series.total).
'Angemeldet'-Zellen werden übersprungen – sie benötigen die Admin-API (get_enrollments).
"""
from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path

_root = Path(__file__).parent.parent
sys.path.insert(0, str(_root))

from dotenv import load_dotenv
load_dotenv(_root / ".env")

from openpyxl import load_workbook

from ausleihe import AusleiheClient, NotFoundError

_HERE = Path(__file__).parent / "Old - Webscraper for Excel"
CONFIG_PATH = _HERE / "config.json"
EXCEL_PATH = _HERE / "Bestand- und Nachbestellungsliste 2025.xlsx"


def parse_entries(entries: list[dict]) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    """Parse config entries into (isbn, cell) pairs split by type.

    Returns:
        bestand_entries:    cells written from Series.total
        angemeldet_entries: cells written from enrollment count (not yet implemented)
    """
    bestand: list[tuple[str, str]] = []
    angemeldet: list[tuple[str, str]] = []
    current_isbn: str | None = None

    for entry in entries:
        if "url" in entry:
            m = re.search(r"/series/(\d+)", entry["url"])
            current_isbn = m.group(1) if m else None

        if current_isbn is None:
            continue

        cell: str = entry["excelCell"]
        xpath: str = entry["elementToRead"]

        if "div[2]/div[2]" in xpath:
            angemeldet.append((current_isbn, cell))
        else:
            bestand.append((current_isbn, cell))

    return bestand, angemeldet


def main() -> None:
    with open(CONFIG_PATH, encoding="utf-8") as f:
        config = json.load(f)

    bestand_entries, angemeldet_entries = parse_entries(config["entries"])
    unique_isbns = {isbn for isbn, _ in bestand_entries}

    print(f"Verbinde mit IServ ({os.environ.get('ISERV_DOMAIN', '?')})...")
    client = AusleiheClient()
    print(f"Lade {len(unique_isbns)} Serien von der API...")

    series_total: dict[str, int] = {}
    series_title: dict[str, str] = {}
    for isbn in sorted(unique_isbns):
        try:
            s = client.series.get_by_isbn(isbn)
            if s.total is not None:
                series_total[isbn] = s.total
                series_title[isbn] = s.title
            else:
                print(f"  WARNUNG: {isbn} — 'total' fehlt in API-Antwort")
        except NotFoundError:
            print(f"  WARNUNG: {isbn} — nicht gefunden (möglicherweise abgelöste Serie)")
        except Exception as e:
            print(f"  FEHLER: {isbn} — {e}")

    wb = load_workbook(str(EXCEL_PATH))
    ws = wb[config["excelSheetName"]]

    changed: list[tuple[str, object, int, str]] = []
    not_found: list[str] = []

    for isbn, cell in bestand_entries:
        if isbn not in series_total:
            not_found.append(f"{cell} (ISBN {isbn})")
            continue
        old = ws[cell].value
        new = series_total[isbn]
        if old != new:
            ws[cell] = new
            changed.append((cell, old, new, series_title[isbn]))

    print()
    if changed:
        wb.save(str(EXCEL_PATH))
        print(f"{len(changed)} Zelle(n) aktualisiert:")
        for cell, old, new, title in changed:
            print(f"  {cell}: {old} -> {new}  [{title}]")
    else:
        print("Keine Änderungen.")

    if not_found:
        print(f"\n{len(not_found)} Bestand-Zelle(n) übersprungen (ISBN nicht in API gefunden):")
        for s in not_found:
            print(f"  {s}")

    if angemeldet_entries:
        cells = ", ".join(cell for _, cell in angemeldet_entries)
        print(f"\nHINWEIS: {len(angemeldet_entries)} 'Angemeldet'-Zellen nicht aktualisiert ({cells}).")
        print("  Benötigt Admin-API (get_enrollments) – noch nicht implementiert.")


if __name__ == "__main__":
    main()
