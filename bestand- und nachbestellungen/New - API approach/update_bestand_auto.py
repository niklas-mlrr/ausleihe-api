#!/usr/bin/env python3
"""Auto-Discovery: befüllt 'Angemeldet'-, 'Bezahlt'-, 'Bestellt'- und 'Bestand'-Zellen
automatisch und trägt Nachbestellbedarf in das Sheet 'zu Bestellen' ein.

Im Unterschied zu update_bestand.py braucht dieses Skript keine config.json-Mappings.
Es liest die Excel-Struktur selbst aus und matched Bücher anhand von Fach-Zeilen.

Algorithmus:
  - Spalte A zeilenweise von oben durchgehen
  - "Fach"-Zeile oder "Zustand"-Zeile → merken
  - "Jahrgang X"-Zeile → ausleihbare Bücher für diesen Jahrgang laden (API GET)
      - Gefiltert: borrowable=True, Leihpreis > 0, kein "eBook" im Titel
  - Für jede Spalte: Fach-Label aus nächster darüber liegender Fach-Zeile bestimmen
      - Klammerzusatz im Fach (z.B. "Politik (eA)") = Hinweis auf Serientitel, nicht Teil des Fachs
      - Falls Fach-Zelle leer → nächst höhere Fach-Zeile als Fallback
  - Zustand-Label aus nächster darüber liegender Zustand-Zeile bestimmen
      - Falls leer → Fallback auf nächst höhere Zustand-Zeile
  - Passt das Buch und ist Zustand "Angemeldet", "Bezahlt", "Bestand" oder "Bestellt" → Wert eintragen
    - "Bestellt": Summe aus Sheet "bestellt" Spalte C für alle Zeilen wo Spalte F (ISBN, "-" ignoriert) passt;
      None wenn ISBN dort nicht vorkommt
  - Bereits bearbeitete Ankerzellen überspringen (Mehrjahresbände / Zellenverbünde)
  - Abbruch nach mehr als 3 nicht identifizierbaren Zeilen in Folge
  - Am Ende: Sheet "zu Bestellen" ab Zeile 2 leeren und mit Büchern befüllen,
    bei denen (Angemeldet - Bestand - Bestellt) > 0 ist.

Verwendung:
    python3 update_bestand_auto.py [--dry-run] [--schoolyear 2025/2026]
                                   [--excel "Bestand- und Nachbestellungsliste 2026.xlsx"]
                                   [--sheet "Bestand- und Nachbestellung"]
                                   [-v | --verbose]

Nur GET-Zugriffe auf die API. Kein Schreiben in die Datenbank.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

_HERE = Path(__file__).parent
_ROOT = _HERE.parent.parent
sys.path.insert(0, str(_ROOT))

from dotenv import load_dotenv
load_dotenv(_ROOT / ".env")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    import isbnlib as _isbnlib
    def format_isbn(isbn: str) -> str:
        try:
            masked = _isbnlib.mask(isbn)
            return masked if masked else isbn
        except Exception:
            return isbn
except ImportError:
    def format_isbn(isbn: str) -> str:  # type: ignore[misc]
        return isbn

from ausleihe import AusleiheClient, NotFoundError


# Excel-Anzeigeformat der "Stand"-Zelle: z.B. "Dienstag, 30.06.2026 17:25:05"
# (TTTT, TT.MM.JJJJ hh:mm:ss). Punkte/Komma/Leerzeichen sind escaped (literal),
# der Wert selbst ist ein echtes datetime → Excel rendert es über dieses Format.
STAND_NUMBER_FORMAT = r"dddd\,\ dd\.mm\.yyyy\ hh:mm:ss"


# ── Excel-Hilfsfunktionen ─────────────────────────────────────────────────────

def resolve_anchor(ws, row: int, col: int) -> tuple[int, int]:
    """Gibt (anchor_row, anchor_col) zurück – bei Zellenverbund die oben-links-Zelle."""
    cell_ref = f"{get_column_letter(col)}{row}"
    for merged in ws.merged_cells.ranges:
        if cell_ref in merged:
            return merged.min_row, merged.min_col
    return row, col


def find_fach_for_col(ws, fach_rows: list[int], col: int) -> str | None:
    """Fach-Label für Spalte col aus der nächsten (untersten) Fach-Zeile mit Inhalt.
    Ist die Zelle leer, wird die nächsthöhere Fach-Zeile als Fallback genommen.
    """
    for fach_row in reversed(fach_rows):
        ar, ac = resolve_anchor(ws, fach_row, col)
        if ar != fach_row:
            # Anker liegt in einer anderen Zeile – diese Zeile überspringen
            continue
        val = ws.cell(ar, ac).value
        if val is not None:
            return str(val)
    return None


def find_zustand_for_col(ws, zustand_rows: list[int], col: int) -> str | None:
    """Zustand-Label für Spalte col, mit Fallback auf höhere Zustand-Zeilen."""
    for zustand_row in reversed(zustand_rows):
        ar, ac = resolve_anchor(ws, zustand_row, col)
        if ar != zustand_row:
            continue
        val = ws.cell(ar, ac).value
        if val is not None:
            return str(val)
    return None


# ── Zeilen-Klassifikation ─────────────────────────────────────────────────────

def classify_row(ws, row: int) -> str:
    """'fach' | 'zustand' | 'stand' | 'jahrgang' | 'other'"""
    val = ws.cell(row, 1).value
    if val == "Fach":
        return "fach"
    if val == "Zustand":
        return "zustand"
    if val == "Stand":
        return "stand"
    if isinstance(val, str) and re.match(r"Jahrgang\s+\d+", val):
        return "jahrgang"
    return "other"


def extract_grade(ws, row: int) -> int | None:
    val = ws.cell(row, 1).value
    m = re.match(r"Jahrgang\s+(\d+)", str(val)) if val else None
    return int(m.group(1)) if m else None


# ── Buch-Hilfsfunktionen ──────────────────────────────────────────────────────

def strip_hint(text: str) -> tuple[str, str | None]:
    """Trennt Serientitel-Hinweis in Klammern vom Fach-Namen.
    'Politik (eA)' → ('Politik', 'eA')
    'Deutsch'      → ('Deutsch', None)
    """
    m = re.match(r"^(.*?)\s*\((.+)\)\s*$", text.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return text.strip(), None


def load_grade_books(client: AusleiheClient, sy_id: str, bl_id: int) -> list[dict]:
    """Alle ausleihbaren Bücher einer Bücherliste (nur GET, gefiltert)."""
    try:
        detail = client.schoolyears.get_booklist(sy_id, bl_id)
    except NotFoundError:
        return []

    seen_isbns: set[str] = set()
    books: list[dict] = []
    for sec in detail.get("sections", []):
        for opt in sec.get("options", []):
            for item in opt.get("items", []):
                if not item.get("borrowable"):
                    continue
                sd = item.get("series_data", {}) or {}
                if (sd.get("fee") or 0) <= 0:
                    continue
                title = sd.get("title", "") or ""
                if "eBook" in title:
                    continue
                isbn = sd.get("isbn") or item.get("series") or ""
                if isbn in seen_isbns:
                    continue
                seen_isbns.add(isbn)
                books.append({
                    "isbn": isbn,
                    "title": title,
                    "subjects": sd.get("subjectsFlat", []) or [],
                })
    return books


# Abkürzungen die nicht literal im Buchtitel stehen, sondern als Langform erscheinen
# Schlüssel lowercase für case-insensitiven Lookup
_HINT_EXPANSIONS: dict[str, str] = {
    "ea": "Erhöhtes",
    "ga": "Grundlegendes",
}

# Linke Wortgrenze: kein vorangehender Buchstabe (inkl. Umlaute)
_LEFT_BOUNDARY = r"(?<![a-zA-ZäöüÄÖÜß])"


def match_book(books: list[dict], subject: str, hint: str | None) -> dict | None:
    """Sucht passendes Buch nach Fach; Klammerzusatz wird case-insensitiv gesucht."""
    candidates = [b for b in books if subject in b["subjects"]]
    if not candidates:
        return None
    if hint:
        terms = [hint]
        expansion = _HINT_EXPANSIONS.get(hint.lower())
        if expansion:
            terms.append(expansion)
        def in_title(title: str) -> bool:
            t = title.lower()
            return any(
                re.search(_LEFT_BOUNDARY + re.escape(term.lower()), t)
                for term in terms
            )
        narrowed = [b for b in candidates if in_title(b["title"])]
        return narrowed[0] if narrowed else None
    return candidates[0]


# ── Anmelde-Zählung per Jahrgang ──────────────────────────────────────────────

def fetch_enrollment_counts_by_grade(
    client: AusleiheClient, sy_id: str
) -> tuple[dict[tuple[int, str], int], dict[tuple[int, str], int]]:
    """Zählt Anmeldungen und Bezahlungen pro (Jahrgang, ISBN).

    enrolled[(grade, isbn)] = Anzahl nicht-gelöschter Anmeldungen, die dieses Buch enthalten
    paid[(grade, isbn)]     = davon mit amountOpen == 0 (vollständig bezahlt)
    """
    enrollments = client.admin.get_enrollments(sy_id)
    enrolled: dict[tuple[int, str], int] = {}
    paid: dict[tuple[int, str], int] = {}

    for enr in enrollments:
        if enr.get("deleted_at"):
            continue
        grade = (enr.get("Booklist") or {}).get("grade")
        if grade is None:
            continue
        is_paid = enr.get("amountOpen", 1) == 0
        for item in enr.get("booklistItems", []):
            isbn = item.get("series")
            if not isbn:
                continue
            key = (grade, isbn)
            enrolled[key] = enrolled.get(key, 0) + 1
            if is_paid:
                paid[key] = paid.get(key, 0) + 1

    return enrolled, paid


def fetch_series_data(client: AusleiheClient) -> dict[str, dict]:
    """Lädt alle Serien; gibt pro ISBN ein Dict mit 'total', 'publisher', 'price'."""
    series_list = client.series.get_all(detailed=True)
    return {
        s.isbn: {
            "total": s.total or 0,
            "publisher": s.publisher or "",
            "price": s.price or 0.0,
            "title": s.title or "",
        }
        for s in series_list
        if s.isbn
    }


def fetch_bestand_by_isbn(client: AusleiheClient) -> dict[str, int]:
    """Gesamtbestand (Exemplaranzahl) pro ISBN aus allen Serien."""
    series_list = client.series.get_all(detailed=True)
    return {s.isbn: (s.total or 0) for s in series_list if s.isbn}


def load_bestellt_counts(ws_bestellt) -> dict[str, int]:
    """Liest Sheet 'bestellt': Spalte F = ISBN (evtl. mit '-'), Spalte C = Stückzahl.
    Gibt pro normierter ISBN (ohne '-') die Summe aller Stückzahlen zurück.
    Zeile 1 (Kopfzeile) wird übersprungen.
    """
    counts: dict[str, int] = {}
    for row in range(2, ws_bestellt.max_row + 1):
        isbn_raw = ws_bestellt.cell(row, 6).value  # Spalte F
        count_raw = ws_bestellt.cell(row, 3).value  # Spalte C
        if isbn_raw is None or count_raw is None:
            continue
        isbn_norm = str(isbn_raw).replace("-", "").strip()
        if not isbn_norm:
            continue
        try:
            count = int(count_raw)
        except (ValueError, TypeError):
            continue
        counts[isbn_norm] = counts.get(isbn_norm, 0) + count
    return counts


# ── Hauptalgorithmus ──────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Auto-Befüllung von Angemeldet-/Bezahlt-Zellen anhand Excel-Struktur"
    )
    parser.add_argument("--dry-run", action="store_true", help="Keine Änderungen speichern")
    parser.add_argument("--schoolyear", help="Schuljahr-ID, z.B. 2025/2026")
    parser.add_argument("--excel", help="Excel-Dateiname (überschreibt config.json)")
    parser.add_argument("--sheet", help="Tabellenblatt-Name (überschreibt config.json)")
    parser.add_argument("-v", "--verbose", action="store_true", help="Detaillierte Debug-Ausgaben")
    args = parser.parse_args()

    # Abfragezeitpunkt (für "Stand"-Zeile) – echtes datetime, damit Excel es über
    # das Zellen-Datumsformat als "TTTT, TT.MM.JJJJ hh:mm:ss" anzeigt.
    abfrage_zeitpunkt = datetime.now()

    print(f"Verbinde mit IServ ({os.environ.get('ISERV_DOMAIN', '?')})...")
    client = AusleiheClient()

    sy_id = args.schoolyear or client.schoolyears.get_current()["id"]
    print(f"Schuljahr: {sy_id}")

    # Excel-Datei und Tabellenblatt bestimmen
    with open(_HERE / "config.json", encoding="utf-8") as f:
        config = json.load(f)
    excel_filename = args.excel or config["excel_file"]
    sheet_name = args.sheet or config["sheet_name"]
    excel_path = _HERE / excel_filename
    print(f"Excel: {excel_path.name}  |  Blatt: {sheet_name}")

    # Bücherlisten für alle Jahrgänge laden
    print("Lade Bücherlisten...")
    booklists = client.schoolyears.get_booklists(sy_id)
    booklists_by_grade = {bl["grade"]: bl for bl in booklists if bl.get("grade") is not None}

    # Anmeldezahlen laden
    print("Lade Anmeldungen...")
    enrolled_counts, paid_counts = fetch_enrollment_counts_by_grade(client, sy_id)

    # Serien-Daten (Bestand, Verlag, Preis) laden
    print("Lade Serien-Daten...")
    series_data = fetch_series_data(client)
    bestand_counts = {isbn: d["total"] for isbn, d in series_data.items()}

    if args.verbose:
        print(f"  enrolled_counts: {len(enrolled_counts)} Einträge, paid_counts: {len(paid_counts)} Einträge")
        print(f"  series_data:     {len(series_data)} Serien")
        sample = list(enrolled_counts.items())[:5]
        for k, v in sample:
            isbn = k[1]
            sd = series_data.get(isbn, {})
            print(f"    {k}: enrolled={v}, paid={paid_counts.get(k, 0)}, bestand={sd.get('total', 0)}")

    wb = load_workbook(str(excel_path))
    ws = wb[sheet_name]

    # "bestellt"-Sheet einlesen
    ws_bestellt = wb["bestellt"]
    bestellt_counts = load_bestellt_counts(ws_bestellt)
    if args.verbose:
        print(f"\nbestellt-Sheet: {len(bestellt_counts)} ISBNs mit Bestellungen")
        for isbn_norm, cnt in bestellt_counts.items():
            print(f"  {isbn_norm}: {cnt}")

    if args.verbose:
        print(f"\nExcel geladen: {ws.max_row} Zeilen, {ws.max_column} Spalten")
        print("Erste Zeilen (Spalte A):")
        for r in range(1, min(ws.max_row + 1, 20)):
            val = ws.cell(r, 1).value
            print(f"  Zeile {r:3d}: A={val!r}")

    fach_rows: list[int] = []      # alle bisher gesehenen Fach-Zeilen (aufsteigend)
    zustand_rows: list[int] = []   # alle bisher gesehenen Zustand-Zeilen (aufsteigend)
    stand_rows: list[int] = []     # alle bisher gesehenen Stand-Zeilen (für Abfragezeitpunkt)
    grade_books_cache: dict[int, list[dict]] = {}
    processed_anchors: set[str] = set()               # Zellenverbund-Dedup (gleiche Zelle)
    processed_bestand_isbns: set[str] = set()          # Bestand: global je ISBN nur einmal
    processed_bestellt_isbns: set[str] = set()         # Bestellt: global je ISBN nur einmal
    processed_enrollment: set[tuple[int, str, str]] = set()  # Angemeldet/Bezahlt: je (grade, isbn, zustand)
    consecutive_other = 0
    changes: list[str] = []

    # Sammelt pro ISBN die geschriebenen Werte für die "zu Bestellen"-Ausgabe.
    # isbn → {"angemeldet": int, "bestand": int, "bestellt": int|None,
    #          "title": str, "grades": set[int], "fach": str}
    zu_bestellen_data: dict[str, dict] = {}

    print("\nAnalysiere Excel-Struktur...\n")

    for row in range(1, ws.max_row + 1):
        row_type = classify_row(ws, row)

        if row_type == "fach":
            fach_rows.append(row)
            consecutive_other = 0
            if args.verbose:
                print(f"  Zeile {row}: FACH erkannt")
            continue

        if row_type == "zustand":
            zustand_rows.append(row)
            consecutive_other = 0
            if args.verbose:
                print(f"  Zeile {row}: ZUSTAND erkannt")
            continue

        if row_type == "stand":
            stand_rows.append(row)
            consecutive_other = 0
            if args.verbose:
                print(f"  Zeile {row}: STAND erkannt")
            continue

        if row_type == "other":
            consecutive_other += 1
            if args.verbose:
                a_val = ws.cell(row, 1).value
                print(f"  Zeile {row}: other (consecutive={consecutive_other}, A={a_val!r})")
            if consecutive_other > 3:
                print(f"Zeile {row}: Abbruch – mehr als 3 nicht erkannte Zeilen in Folge.")
                break
            continue

        # Jahrgang-Zeile
        consecutive_other = 0
        grade = extract_grade(ws, row)
        if args.verbose:
            print(f"\n  Zeile {row}: JAHRGANG {grade} | fach_rows={fach_rows} | zustand_rows={zustand_rows}")
        if grade is None or not fach_rows:
            if args.verbose:
                print(f"    -> Übersprungen (grade={grade}, fach_rows leer={not fach_rows})")
            continue

        # Bücher für diesen Jahrgang laden (gecacht)
        if grade not in grade_books_cache:
            bl = booklists_by_grade.get(grade)
            if bl:
                print(f"  Lade Bücherliste Jahrgang {grade}...")
                grade_books_cache[grade] = load_grade_books(client, sy_id, bl["id"])
            else:
                print(f"  WARNUNG: Keine Bücherliste für Jahrgang {grade} im Schuljahr {sy_id}")
                grade_books_cache[grade] = []
        books = grade_books_cache[grade]

        if args.verbose:
            print(f"    {len(books)} Bücher geladen")
            for b in books:
                print(f"      isbn={b['isbn']} subjects={b['subjects']} title={b['title'][:50]!r}")

        # Alle Spalten der Jahrgang-Zeile durchsuchen
        for col in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col)

            # Zustand-Label für diese Spalte bestimmen
            zustand_label = find_zustand_for_col(ws, zustand_rows, col)
            if zustand_label is None:
                if args.verbose:
                    print(f"    Sp.{col_letter}: kein Zustand-Label → skip")
                continue
            zustand_norm = zustand_label.strip().lower()
            if zustand_norm not in ("angemeldet", "bezahlt", "bestand", "bestellt"):
                if args.verbose:
                    print(f"    Sp.{col_letter}: Zustand={zustand_label!r} (nicht angemeldet/bezahlt/bestand/bestellt) → skip")
                continue

            # Fach-Label für diese Spalte bestimmen (mit Fallback auf höhere Fach-Zeilen)
            fach_val = find_fach_for_col(ws, fach_rows, col)
            if fach_val is None:
                if args.verbose:
                    print(f"    Sp.{col_letter}: [{zustand_label}] kein Fach-Label → skip")
                continue

            subject, hint = strip_hint(fach_val)
            book = match_book(books, subject, hint)
            if book is None:
                if args.verbose:
                    print(f"    Sp.{col_letter}: [{zustand_label}] Fach={fach_val!r} → kein Buch-Match (subjects in Bücherliste: {[b['subjects'] for b in books]})")
                continue

            # Ankerzelle der Jahrgang-Zeile bestimmen (Zellenverbund-Auflösung)
            ar, ac = resolve_anchor(ws, row, col)
            anchor_ref = f"{get_column_letter(ac)}{ar}"

            if anchor_ref in processed_anchors:
                if args.verbose:
                    print(f"    Sp.{col_letter}: {anchor_ref} bereits verarbeitet (Zellenverbund) → skip")
                continue
            processed_anchors.add(anchor_ref)

            isbn = book["isbn"]
            if zustand_norm == "bestand":
                if isbn in processed_bestand_isbns:
                    if args.verbose:
                        print(f"    Sp.{col_letter}: isbn={isbn}/Bestand bereits eingetragen → skip")
                    continue
                processed_bestand_isbns.add(isbn)
            elif zustand_norm == "bestellt":
                if isbn in processed_bestellt_isbns:
                    if args.verbose:
                        print(f"    Sp.{col_letter}: isbn={isbn}/Bestellt bereits eingetragen → skip")
                    continue
                processed_bestellt_isbns.add(isbn)
            else:
                enr_key = (grade, isbn, zustand_norm)
                if enr_key in processed_enrollment:
                    if args.verbose:
                        print(f"    Sp.{col_letter}: isbn={isbn}/{zustand_label} Jg.{grade} bereits eingetragen → skip")
                    continue
                processed_enrollment.add(enr_key)

            key = (grade, isbn)
            if zustand_norm == "angemeldet":
                new_val: int | None = enrolled_counts.get(key, 0)
            elif zustand_norm == "bezahlt":
                new_val = paid_counts.get(key, 0)
            elif zustand_norm == "bestand":
                new_val = bestand_counts.get(isbn, 0)
            else:  # bestellt
                isbn_norm = isbn.replace("-", "")
                new_val = bestellt_counts.get(isbn_norm) if isbn_norm in bestellt_counts else None

            # Tracking für "zu Bestellen"-Sheet
            entry = zu_bestellen_data.setdefault(isbn, {
                "angemeldet": 0,
                "bestand": 0,
                "bestellt": None,
                "title": book["title"],
                "grades": set(),
                "fach": fach_val,
            })
            entry["grades"].add(grade)
            if zustand_norm == "angemeldet" and new_val is not None:
                entry["angemeldet"] = (entry["angemeldet"] or 0) + new_val
            elif zustand_norm == "bestand" and new_val is not None:
                entry["bestand"] = new_val
            elif zustand_norm == "bestellt":
                entry["bestellt"] = new_val

            if args.verbose:
                isbn_norm = isbn.replace("-", "")
                print(
                    f"    Sp.{col_letter}: {anchor_ref} {fach_val!r}/{zustand_label} → "
                    f"isbn={isbn}, enrolled={enrolled_counts.get(key,'–')}, "
                    f"paid={paid_counts.get(key,'–')}, bestand={bestand_counts.get(isbn,'–')}, "
                    f"bestellt_sheet={bestellt_counts.get(isbn_norm,'–')}, new_val={new_val}"
                )

            old_val = ws[anchor_ref].value
            ws[anchor_ref].value = new_val

            hint_str = f" ({hint})" if hint else ""
            changes.append(
                f"  {anchor_ref}: {old_val!r} -> {new_val!r}"
                f"  [{subject}{hint_str} Jg.{grade}, {book['title']}, {zustand_label}]"
            )

    # ── Abfragezeitpunkt in "Stand"-Zeile(n) eintragen ───────────────────────
    # In Spalte B jeder erkannten Stand-Zeile, Format TT.MM.JJJJ hh:mm:ss
    for stand_row in stand_rows:
        ar, ac = resolve_anchor(ws, stand_row, 2)  # Spalte B
        cell = ws.cell(ar, ac)
        anchor_ref = f"{get_column_letter(ac)}{ar}"
        old_val = cell.value
        cell.value = abfrage_zeitpunkt              # echtes datetime (Excel-Datumswert)
        cell.number_format = STAND_NUMBER_FORMAT    # Anzeige: TTTT, TT.MM.JJJJ hh:mm:ss
        changes.append(
            f"  {anchor_ref}: {old_val!r} -> "
            f"{abfrage_zeitpunkt.strftime('%d.%m.%Y %H:%M:%S')}  [Stand/Abfragezeitpunkt]"
        )

    # ── Sheet "zu Bestellen" befüllen ────────────────────────────────────────
    ws_zu = wb["zu Bestellen"]

    # Spalten: B=Jahrgänge, C=Fach, D=Stückzahl, E=Titel, F=Verlag, G=ISBN, H=Neupreis
    zu_bestellen_rows: list[tuple] = []
    for isbn, entry in zu_bestellen_data.items():
        angemeldet = entry["angemeldet"] or 0
        bestand = entry["bestand"] or 0
        bestellt = entry["bestellt"] or 0
        zu_bestellen = angemeldet - bestand - bestellt
        if zu_bestellen > 0:
            sd = series_data.get(isbn, {})
            publisher = sd.get("publisher", "")
            price = sd.get("price", 0.0)
            title = sd.get("title") or entry["title"]
            isbn_fmt = format_isbn(isbn)
            grades_str = min(entry["grades"])
            fach = entry["fach"]
            zu_bestellen_rows.append(
                (grades_str, fach, zu_bestellen + 5, title, publisher, isbn_fmt, price)
            )

    zu_bestellen_rows.sort(key=lambda r: r[3])  # alphabetisch nach Titel

    # Alte Einträge ab Zeile 2 löschen (Spalten B–H)
    for row in range(2, ws_zu.max_row + 1):
        for col in range(2, 9):  # B–H
            ws_zu.cell(row, col).value = None

    # Neue Einträge schreiben
    for i, (grades_str, fach, stueckzahl, title, publisher, isbn_fmt, price) in enumerate(zu_bestellen_rows):
        row = 2 + i
        ws_zu.cell(row, 2).value = grades_str   # B: Jahrgänge
        ws_zu.cell(row, 3).value = fach          # C: Fach
        ws_zu.cell(row, 4).value = stueckzahl    # D: Stückzahl
        ws_zu.cell(row, 5).value = title         # E: Titel
        ws_zu.cell(row, 6).value = publisher     # F: Verlag
        ws_zu.cell(row, 7).value = isbn_fmt      # G: ISBN
        ws_zu.cell(row, 8).value = price         # H: Einzelpreis

    print(f"\n{len(zu_bestellen_rows)} Bücher mit Nachbestellbedarf:")
    for grades_str, fach, stueckzahl, title, publisher, isbn_fmt, price in zu_bestellen_rows:
        print(f"  Jg.{grades_str:2d} [{fach}] +5 → {stueckzahl} Stk.  {title[:45]}  [{isbn_fmt}]")

    # ── Ausgabe & Speichern ──────────────────────────────────────────────────
    print()
    if args.dry_run:
        print("-- DRY RUN: keine Datei wird gespeichert --")

    if changes:
        print(f"{len(changes)} Zelle(n) {'würden aktualisiert' if args.dry_run else 'aktualisiert'}:")
        for c in changes:
            print(c)
    else:
        print("Keine Änderungen.")

    if not args.dry_run:
        wb.save(str(excel_path))
        print(f"\nGespeichert: {excel_path}")


if __name__ == "__main__":
    main()
